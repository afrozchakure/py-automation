import openpyxl as xl
from openpyxl.chart import BarChart, Reference  # Adding Charts to excel

def process_workbook(filename):

    wb = xl.load_workbook('filename')
    sheet = wb['Sheet1']  # Access the sheet
    # cell = sheet['a1']  # Give the coordinate
    # cell = sheet.cell(1,1)  # Access a cell

    # print(cell.value)  # Get value of cell at specified position
    # print(sheet.max_row)  # Get the maximum row filed

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)  # Get access to cell at row and column = 3
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)  # setting the corrected value
        corrected_price_cell.value = corrected_price  # setting corrected_value at corrected_prince_cell

    values = Reference(sheet,  # All the values in the 4th column
              min_row= 2,
              max_row = sheet.max_row,
              min_col = 4,
              max_col = 4)

    chart = BarChart()  # Creating a BarChart object from the class
    chart.add_data(values)
    sheet.add_chart(chart, 'E2') # Adding the chart to our sheet (E2 in capital)

    wb.save("filename")  # Workbook.save (overwrite the same file)

